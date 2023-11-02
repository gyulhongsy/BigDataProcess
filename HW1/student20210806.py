#!/usr/bin/python3

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

fileName = "student.xlsx"
wb = load_workbook(filename=fileName, data_only=True)
ws = wb['Sheet1']

grade_list = []

#total 계산
for row in range(2, 76):
    mid = ws.cell(row=row, column=3).value
    fin = ws.cell(row=row, column=4).value
    hom = ws.cell(row=row, column=5).value
    att = ws.cell(row=row, column=6).value
    total = round((mid * 0.3) + (fin * 0.35) + (hom * 0.34) + att, 2)
    ws.cell(row=row, column=7, value = total)
    grade_list.append(total)
wb.save(filename=fileName)

grade_list.sort(reverse=True)
ap = int(len(grade_list) * 0.3 * 0.5)
a = int(len(grade_list) * 0.3 * 0.5) + ap
bp = int(len(grade_list) * 0.4 * 0.5) + a
b = int(len(grade_list) * 0.4 * 0.5) + bp
cp = int(len(grade_list) * 0.3 * 0.5) + b

ap_list = grade_list[:ap]
a_list = grade_list[ap:a]
bp_list = grade_list[a:bp]
b_list = grade_list[bp:b]
cp_list = grade_list[b:cp]
c_list = grade_list[cp:]

#학점 부여
grade = ""
for row in range(2, 76):
    total = float(ws.cell(row=row, column=7).value)
    if total < 40.0:
        grade = "F"
    else: 
        for data in ap_list:
            if total == data:
                grade = "A+"
                break
        for data in a_list:
            if total == data:
                grade = "A"
                break
        for data in bp_list:
            if total == data:
                grade = "B+"
                break
        for data in b_list:
            if total == data:
                grade = "B"
                break
        for data in cp_list:
            if total == data:
                grade = "C+"
                break
        for data in c_list:
            if total == data:
                grade = "C"
                break
    ws.cell(row=row, column=8, value=grade)
wb.save(filename=fileName)